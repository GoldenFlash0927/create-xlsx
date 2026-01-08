#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel文件更新脚本
根据模板Excel文件，生成新的Excel文件，更新考勤月份和日期信息
"""

import os
import openpyxl
from datetime import datetime, date
from openpyxl.utils import get_column_letter
import calendar


def get_current_year_month():
    """获取当前年月"""
    now = datetime.now()
    return now.year, now.month


def get_days_in_month(year, month):
    """获取指定年月的天数"""
    return calendar.monthrange(year, month)[1]


def find_template_file(directory='.'):
    """查找当前目录下的.xlsx模板文件"""
    xlsx_files = [f for f in os.listdir(directory) if f.endswith('.xlsx')]
    if not xlsx_files:
        raise FileNotFoundError("当前目录下未找到.xlsx文件")
    
    # 优先查找包含"模版"或"模板"的文件
    for filename in xlsx_files:
        if '模版' in filename or '模板' in filename:
            return filename
    
    # 如果没有找到模板文件，使用最新的文件（排除当前月份的文件）
    current_year, current_month = get_current_year_month()
    current_filename = f"{current_year}年{current_month}月.xlsx"
    
    # 排除当前月份的文件，使用其他文件作为模板
    template_files = [f for f in xlsx_files if f != current_filename]
    if template_files:
        # 按修改时间排序，使用最新的作为模板
        template_files.sort(key=lambda f: os.path.getmtime(os.path.join(directory, f)), reverse=True)
        return template_files[0]
    
    # 如果只有当前月份的文件，也使用它
    return xlsx_files[0]


def update_attendance_sheet(worksheet, year, month):
    """更新考勤工作表"""
    # 更新年份和月份（第2行）
    worksheet.cell(2, 2).value = year  # B2: 年份
    worksheet.cell(2, 5).value = month  # E2: 月份
    
    # 获取该月的天数
    days_in_month = get_days_in_month(year, month)
    
    # 更新第5行的日期（从B5开始）
    # 只更新日期值，保留原有的格式和数据验证规则
    for day in range(1, days_in_month + 1):
        col = day + 1  # B列是第2列，所以day=1对应B列(2)，day=2对应C列(3)...
        current_date = date(year, month, day)
        excel_date = openpyxl.utils.datetime.to_excel(current_date)
        cell = worksheet.cell(5, col)
        # 只更新值，保留原有的格式
        original_format = cell.number_format
        cell.value = excel_date
        # 如果原格式不是日期格式，则设置为日期格式
        if original_format == 'General' or '月' not in original_format:
            cell.number_format = 'm"月"d"日"'
        else:
            cell.number_format = original_format
    
    # 如果该月天数少于31天，清空多余的日期单元格（但保留格式）
    for day in range(days_in_month + 1, 32):
        col = day + 1
        cell = worksheet.cell(5, col)
        original_format = cell.number_format
        cell.value = None
        cell.number_format = original_format


def update_sheet_name(workbook, old_name_pattern, new_name):
    """更新工作表名称"""
    for sheet_name in workbook.sheetnames:
        if old_name_pattern in sheet_name:
            # 提取"手工提成"后面的月份数字，替换为新的月份
            if "手工提成" in sheet_name:
                workbook[sheet_name].title = new_name
                break


def generate_excel_file(template_file, output_file, year, month):
    """生成新的Excel文件"""
    # 加载模板文件，保留所有内容（包括数据验证、格式等）
    print(f"正在加载模板文件: {template_file}")
    # keep_vba=True 保留VBA代码（如果有）
    # data_only=False 保留公式而不是计算结果
    workbook = openpyxl.load_workbook(template_file, keep_vba=True, data_only=False)
    
    # 更新考勤工作表
    if '考勤' in workbook.sheetnames:
        print(f"正在更新'考勤'工作表...")
        attendance_sheet = workbook['考勤']
        
        # 检查并打印数据验证规则（用于调试）
        if attendance_sheet.data_validations.dataValidation:
            print(f"  检测到 {len(attendance_sheet.data_validations.dataValidation)} 个数据验证规则")
            for dv in attendance_sheet.data_validations.dataValidation:
                print(f"    - 范围: {dv.ranges}, 选项: {dv.formula1 if hasattr(dv, 'formula1') else '无'}")
        
        update_attendance_sheet(attendance_sheet, year, month)
        print(f"✓ 已更新考勤月份为 {year}年{month}月")
    else:
        print("⚠ 警告: 未找到'考勤'工作表")
    
    # 更新手工提成工作表名称
    month_str = f"{month}月"
    for sheet_name in workbook.sheetnames:
        if "手工提成" in sheet_name and any(char.isdigit() for char in sheet_name):
            new_name = f"手工提成{month_str}"
            workbook[sheet_name].title = new_name
            print(f"✓ 已更新工作表名称: '{sheet_name}' -> '{new_name}'")
            break
    
    # 保存新文件，确保保留所有内容
    print(f"正在保存文件: {output_file}")
    workbook.save(output_file)
    
    # 验证保存的文件是否保留了数据验证
    verify_workbook = openpyxl.load_workbook(output_file)
    if '考勤' in verify_workbook.sheetnames:
        verify_sheet = verify_workbook['考勤']
        if verify_sheet.data_validations.dataValidation:
            print(f"✓ 已确认数据验证规则已保留（{len(verify_sheet.data_validations.dataValidation)} 个规则）")
        else:
            print("⚠ 警告: 数据验证规则可能未正确保留")
    
    print(f"✓ 文件已成功生成: {output_file}")


def main():
    """主函数"""
    try:
        # 获取当前年月
        year, month = get_current_year_month()
        print(f"当前年月: {year}年{month}月")
        
        # 查找模板文件
        template_file = find_template_file()
        print(f"找到模板文件: {template_file}")
        
        # 生成输出文件名
        output_file = f"{year}年{month}月.xlsx"
        
        # 如果输出文件已存在，询问是否覆盖
        if os.path.exists(output_file):
            response = input(f"文件 {output_file} 已存在，是否覆盖？(y/n): ")
            if response.lower() != 'y':
                print("操作已取消")
                return
        
        # 生成新文件
        generate_excel_file(template_file, output_file, year, month)
        print("\n✅ 所有操作完成！")
        
    except Exception as e:
        print(f"❌ 发生错误: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == '__main__':
    main()

